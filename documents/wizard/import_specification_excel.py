import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ImportSpecificationExcel(models.TransientModel):
    _name = 'dino.import.specification.excel'
    _description = 'Import Specification from Excel'

    file_data = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')
    parser_type = fields.Selection([
        ('sapriz', 'Sapriz Invoice (1C)'),
        ('universal', 'Universal Template'),
    ], string='File Format', default='sapriz', required=True)

    def _clean_text(self, text):
        """Очистка текста: убрать лишние пробелы и пробелы вокруг дефисов"""
        if not text:
            return text
        
        import re
        
        # Преобразовать в строку
        text = str(text).strip()
        
        # Заменить множественные пробелы на одинарные
        text = re.sub(r'\s+', ' ', text)
        
        # Убрать пробелы вокруг дефисов: " - " -> "-"
        text = re.sub(r'\s*-\s*', '-', text)
        
        return text

    def action_import(self):
        """Парсинг Excel и создание строк спецификации"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(_("Python library 'openpyxl' is not installed. Please install it to use Excel import."))

        # Получить активный документ
        document_id = self.env.context.get('active_id')
        if not document_id:
            raise UserError(_("No active document found."))
        
        document = self.env['dino.operation.document'].browse(document_id)
        if not document.exists():
            raise UserError(_("Document not found."))

        # Декодировать файл
        try:
            file_content = base64.b64decode(self.file_data)
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Cannot read Excel file. Error: %s") % str(e))

        # Найти заголовок и данные для Excel в стиле 1С с множеством пустых колонок
        # Стратегия: найти строку с ключевыми заголовками, затем найти строки данных
        # Также извлечь информацию о документе перед заголовком
        header_row_idx = None
        data_start_row = None
        doc_number = None
        doc_date = None
        
        # Первый проход: найти строку заголовка и извлечь информацию о документе перед ней
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not row:
                continue
            
            # Получить все непустые ячейки в нижнем регистре
            row_values = [str(cell).lower().strip() if cell else '' for cell in row]
            row_text = ' '.join(row_values)
            
            # Искать информацию о документе ДО нахождения заголовка
            if not header_row_idx:
                # Проверить все ячейки на наличие номера и даты документа
                for cell in row:
                    if cell and isinstance(cell, str):
                        cell_lower = cell.lower()
                        
                        # Искать "Рахунок на оплату № XXX від DD month YYYY"
                        if 'рахунок' in cell_lower and '№' in cell:
                            import re
                            from datetime import datetime
                            
                            # Извлечь номер после №
                            number_match = re.search(r'№\s*([^\s]+)', cell)
                            if number_match:
                                doc_number = number_match.group(1).strip()
                            
                            # Извлечь дату - украинские названия месяцев
                            months_ua = {
                                'січня': 1, 'лютого': 2, 'березня': 3, 'квітня': 4,
                                'травня': 5, 'червня': 6, 'липня': 7, 'серпня': 8,
                                'вересня': 9, 'жовтня': 10, 'листопада': 11, 'грудня': 12
                            }
                            
                            for month_name, month_num in months_ua.items():
                                if month_name in cell_lower:
                                    # Паттерн: "DD month_name YYYY"
                                    date_match = re.search(rf'(\d{{1,2}})\s+{month_name}\s+(\d{{4}})', cell_lower)
                                    if date_match:
                                        day = int(date_match.group(1))
                                        year = int(date_match.group(2))
                                        try:
                                            doc_date = datetime(year, month_num, day).date()
                                        except:
                                            pass
                                    break
            
            # Искать ключевые слова заголовка
            if ('товар' in row_text or 'послуг' in row_text) and ('кіл' in row_text or 'кол' in row_text):
                header_row_idx = row_idx
                break
        
        # Если заголовок не найден, начать с строки 1
        if not header_row_idx:
            header_row_idx = 1
        
        # Второй проход: найти первую строку данных (где первая непустая ячейка после заголовка - "1")
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            if not row:
                continue
            
            # Найти первую непустую ячейку
            for cell in row:
                if cell and str(cell).strip():
                    try:
                        if str(cell).strip() == '1':
                            data_start_row = row_idx
                            break
                    except:
                        pass
            
            if data_start_row:
                break
        
        if not data_start_row:
            raise UserError(_("Cannot find data rows. Make sure your Excel has a table with row numbers starting from 1."))

        # Получить максимальную последовательность для правильной сортировки
        max_sequence = 0
        if document.specification_ids:
            max_sequence = max(document.specification_ids.mapped('sequence'))

        # Удалить существующие строки если требуется
        if self.env.context.get('replace_existing', False):
            document.specification_ids.unlink()
            max_sequence = 0

        lines_to_create = []
        
        # Получить контрагента документа для работы со справочником
        partner = document.partner_id
        if not partner:
            raise UserError(_("Document must have a partner to import nomenclature"))
        
        # Парсинг строк данных - извлечение непустых ячеек по порядку
        # Ожидаемый порядок: № | Название | Кол-во | Ед.изм. | Цена | ...
        for row_idx in range(data_start_row, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            if not row:
                continue

            # Извлечь непустые ячейки
            cells = [cell for cell in row if cell is not None and str(cell).strip()]
            
            if len(cells) < 4:  # Нужно минимум: №, название, кол-во, цена
                continue
            
            # Первая ячейка должна быть номером строки
            try:
                row_num = str(cells[0]).strip()
                if not row_num.replace('.', '', 1).replace(',', '', 1).isdigit():
                    # Не число - возможно итоги или конец данных
                    break
            except:
                continue
            
            # Вторая ячейка - название
            name = str(cells[1]).strip() if len(cells) > 1 else ''
            if not name or len(name) < 3:
                continue
            
            # Очистить название
            name = self._clean_text(name)
            
            # Третья ячейка - количество
            try:
                qty = float(str(cells[2]).replace(',', '.')) if len(cells) > 2 else 1.0
            except (ValueError, TypeError):
                qty = 1.0
            
            # Пропустить единицу измерения (cells[3])
            
            # Пятая ячейка - цена (индекс 4)
            try:
                price_untaxed = float(str(cells[4]).replace(',', '.')) if len(cells) > 4 else 0.0
            except (ValueError, TypeError):
                price_untaxed = 0.0
            
            # Рассчитать цену с НДС на основе ставки НДС документа
            vat_rate = document.vat_rate or 0.0
            price_with_tax = price_untaxed * (1 + vat_rate / 100)

            # Увеличить последовательность
            max_sequence += 10

            # === ЛОГИКА РАБОТЫ СО СПРАВОЧНИКОМ КОНТРАГЕНТА ===
            # Шаг 1: Найти или создать запись в справочнике контрагента
            # - Если название уже есть → используем существующую запись
            # - Если название новое → создаем новую запись
            supplier_nomenclature = self.env['dino.partner.nomenclature'].find_or_create(
                partner_id=partner.id,
                supplier_name=name,
                auto_create=True
            )
            
            # Шаг 2: Получить связанную складскую номенклатуру (если связь уже настроена)
            # - Если связь есть → показываем её в документе
            # - Если связи нет → оставляем пустым (пользователь настроит позже)
            nomenclature_id = supplier_nomenclature.nomenclature_id.id if supplier_nomenclature.nomenclature_id else False

            # Создать данные строки спецификации
            # supplier_nomenclature_id - ID записи из справочника контрагента (всегда заполнен)
            # nomenclature_id - ID складской номенклатуры (заполнен только если связь уже настроена)
            vals = {
                'document_id': document.id,
                'name': name,
                'quantity': qty,
                'price_untaxed': price_untaxed,
                'price_tax': price_with_tax,
                'sequence': max_sequence,
                'supplier_nomenclature_id': supplier_nomenclature.id,  # ← Запись из справочника контрагента
                'nomenclature_id': nomenclature_id,  # ← Складская номенклатура (может быть False)
            }

            lines_to_create.append(vals)

        # Создать строки
        if lines_to_create:
            self.env['dino.operation.document.specification'].create(lines_to_create)
            
            # Обновить номер и дату документа, если найдены
            update_vals = {}
            if doc_number and not document.number:
                update_vals['number'] = doc_number
            if doc_date and not document.date:
                update_vals['date'] = doc_date
            
            if update_vals:
                document.write(update_vals)
            
            # Закрыть визард и показать уведомление
            message = _('%s lines imported successfully.') % len(lines_to_create)
            if doc_number:
                message += _(' Document number: %s') % doc_number
            if doc_date:
                message += _(' Date: %s') % doc_date.strftime('%d.%m.%Y')
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': message,
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                }
            }
        else:
            raise UserError(_("No valid data found in Excel file."))
